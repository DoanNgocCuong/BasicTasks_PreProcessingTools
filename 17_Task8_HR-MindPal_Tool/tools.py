from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
import time
import openpyxl
import os

def scrape_topcv():
    driver = webdriver.Edge()
    wait = WebDriverWait(driver, 20)

    # Load the workbook or create a new one if it doesn't exist
    try:
        wb = openpyxl.load_workbook('topcv_jobs.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    
    ws = wb.active

    # Add headers to the Excel sheet
    headers = ['Thời gian đăng', 'Kinh nghiệm', 'Mức lương', 'Tên công ty', 'Quy mô công ty', 'Địa chỉ công ty', 'Trang công ty']
    ws.append(headers)  # Add headers to the first row


    try:
        driver.get("https://www.topcv.vn/tim-viec-lam-ai-engineer")

        job_listings = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "job-item-search-result")))
       
        print(f"Số lượng công việc tìm thấy: {len(job_listings)}")
       
        for index, job in enumerate(job_listings, 1):
            try:
                print(f"Processing job {index}/{len(job_listings)}")
                posting_time = job.find_element(By.CSS_SELECTOR, "label.address.mobile-hidden.label-update").get_attribute("data-original-title")
                posting_time = posting_time.replace("Cập nhật ", "")
                print(f"Thời gian đăng: {posting_time}")
                ws[f'A{index}'] = posting_time  # Fill posting time into Excel
                
                # Attempt to click the quick view element
                try:
                    quick_view_element = job.find_element(By.CLASS_NAME, "quick-view-job-detail")
                    driver.execute_script("arguments[0].click();", quick_view_element)
                except NoSuchElementException:
                    print("Không tìm thấy nút xem nhanh.")
        
                # Trích xuất thông tin lương và kinh nghiệm
                try:
                    # Đợi cho đến khi phần tử chứa thông tin lương và kinh nghiệm xuất hiện
                    info_header = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.box-info-header")))
                    
                    # Trích xuất thông tin lương
                    try:
                        salary = info_header.find_element(By.CSS_SELECTOR, "div.box-item-header div.box-item-value[data-original-title$='triệu']").get_attribute("data-original-title")
                        print(f"Mức lương: {salary}")
                        ws[f'C{index+1}'] = salary
                    except NoSuchElementException:
                        print("Không tìm thấy thông tin về mức lương.")
                        ws[f'C{index+1}'] = "Không có thông tin"

                    # Trích xuất thông tin kinh nghiệm
                    try:
                        experience = info_header.find_element(By.CSS_SELECTOR, "div.box-item-header div.box-item-value[data-original-title$='năm']").get_attribute("data-original-title")
                        print(f"Kinh nghiệm: {experience}")
                        ws[f'B{index+1}'] = experience
                    except NoSuchElementException:
                        print("Không tìm thấy thông tin về kinh nghiệm.")
                        ws[f'B{index+1}'] = "Không có thông tin"
                        
        
                for section in ['Mô tả công việc', 'Yêu cầu ứng viên', 'Quyền lợi', 'Địa điểm làm việc', 'Thời gian làm việc']:
                    try:
                        content = wait.until(EC.presence_of_element_located((By.XPATH, f"//h3[contains(text(), '{section}')]/following-sibling::div[1]"))).text
                        print(f"{section}: {content}")
                        ws[f'C{index}'] = content  # Fill content into Excel
                    except:
                        print(f"Không tìm thấy thông tin về {section}.")
                        ws[f'C{index}'] = "Không có thông tin"  # Fill default value into Excel
                

                # Extract company information
                try:
                    # Update selectors based on the new HTML structure
                    company_name = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.job-detail__company--information .company-name a.name"))).text
                    print(f"Tên công ty: {company_name}")
                    ws[f'D{index}'] = company_name  # Fill company name into Excel
 
                    company_scale = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.job-detail__company--information .company-scale .company-value"))).text
                    print(f"Quy mô công ty: {company_scale}")
                    ws[f'E{index}'] = company_scale  # Fill company scale into Excel
 
                    company_address = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.job-detail__company--information .company-address .company-value"))).text
                    print(f"Địa chỉ công ty: {company_address}")
                    ws[f'F{index}'] = company_address  # Fill company address into Excel
 
                    company_link = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.job-detail__company--link a"))).get_attribute("href")
                    print(f"Trang công ty: {company_link}")
                    ws[f'G{index}'] = company_link  # Fill company link into Excel
                except NoSuchElementException as e:
                    print(f"Không thể trích xuất thông tin công ty: Không tìm thấy phần tử - {e}")
                    ws[f'D{index}'] = "Không có thông tin"  # Fill default value into Excel
                except TimeoutException as e:
                    print(f"Không thể trích xuất thông tin công ty: Thời gian chờ đã hết - {e}")
                    ws[f'D{index}'] = "Không có thông tin"  # Fill default value into Excel
                except Exception as e:
                    print(f"Không thể trích xuất thông tin công ty: {e}")
                    ws[f'D{index}'] = "Không có thông tin"  # Fill default value into Excel
                    
                close_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button.btn-close-job-detail")))
                driver.execute_script("arguments[0].click();", close_button)  # Ensure the close button is clicked
                        
            except (NoSuchElementException, TimeoutException) as e:
                print(f"Error processing job {index}: {str(e)}")
                ws[f'A{index}'] = "Không có thông tin"  # Fill default value into Excel

            print("---")

    except WebDriverException:
        print("Trình duyệt đã đóng, đang khởi động lại...")
        driver = webdriver.Edge()
        wait = WebDriverWait(driver, 20)
        # Tiếp tục từ đây
    except Exception as e:
        print(f"Đã xảy ra lỗi: {e}")
    finally:
        driver.quit()
        try:
            wb.save('topcv_jobs.xlsx')  # Attempt to save the workbook
        except PermissionError:
            print("Permission denied: The file is open. Please close it and try again.")
if __name__ == "__main__":
    scrape_topcv()
